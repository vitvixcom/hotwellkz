#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Генератор SEO-лендингов HotWell.kz под [город] + [услугу/ключевые слова].

ПОДХОД: берёт ПОЛНУЮ главную страницу (site/index.html) со ВСЕМИ блоками
(видео-отзывы, отзывы, до/после, технология, бригады, калькулятор, построенные
дома, каталог, что входит, статистика, этапы, о компании, блог, контакты и т.д.)
и адаптирует только SEO-сигналы под город+услугу: <title>, meta description,
keywords, canonical, Open Graph/Twitter, H1, надзаголовок и лид в hero, добавляет
JSON-LD Service + BreadcrumbList и метку источника в форму заявки.

Страница кладётся в КОРЕНЬ сайта (site/<slug>.html), поэтому все относительные
пути (css, js, картинки, ссылки) работают как на главной — без переписывания.
URL получается чистым: https://hotwellkz.kz/<slug>.html

Тексты-заголовки пишет OpenAI как конверсионный копирайтер (цель ~20% заявок).

Запуск:
    python3 generate.py --city "Астана" --service "каркасные дома"
    python3 generate.py --city "Шымкент" --service "дома из СИП-панелей" --kw "цена, под ключ"
    python3 generate.py --city "Алматы" --service "домокомплекты" --no-ai
"""
import os, sys, json, ssl, re, argparse, urllib.request, urllib.error, urllib.parse

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.abspath(os.path.join(HERE, "..", "..", ".."))
SITE = os.path.join(ROOT, "site")
INDEX = os.path.join(SITE, "index.html")
SITEMAP = os.path.join(SITE, "sitemap.xml")
PUBLISHED = os.path.join(HERE, "published.json")
KEYWORDS_XLSX = os.path.join(ROOT, "Ключевые слова.xlsx")
CA = "/root/.ccr/ca-bundle.crt"
CTX = ssl.create_default_context(cafile=CA if os.path.exists(CA) else None)

BASE_URL = os.environ.get("SITE_BASE_URL", "https://hotwellkz.kz").rstrip("/")
TEXT_MODEL = os.environ.get("OPENAI_TEXT_MODEL", "gpt-4o")

TRANSLIT = {'а':'a','б':'b','в':'v','г':'g','д':'d','е':'e','ё':'e','ж':'zh','з':'z',
 'и':'i','й':'y','к':'k','л':'l','м':'m','н':'n','о':'o','п':'p','р':'r','с':'s','т':'t',
 'у':'u','ф':'f','х':'h','ц':'ts','ч':'ch','ш':'sh','щ':'sch','ъ':'','ы':'y','ь':'',
 'э':'e','ю':'yu','я':'ya'}


def esc(s):
    import html as _h
    return _h.escape(str(s or ""), quote=True)


def slugify(text):
    text = (text or "").lower().strip()
    out = []
    for ch in text:
        if ch in TRANSLIT: out.append(TRANSLIT[ch])
        elif ch.isalnum() and ord(ch) < 128: out.append(ch)
        elif ch in " -_/": out.append("-")
    return (re.sub(r"-+", "-", "".join(out)).strip("-") or "lp")[:70]


def _norm(s):
    """Нормализация для сопоставления: нижний регистр, ё→е, дефис/пунктуация → пробел."""
    s = (s or "").lower().replace("ё", "е")
    return re.sub(r"\s+", " ", re.sub(r"[^a-zа-я0-9]+", " ", s)).strip()


# Категории услуг HotWell.kz (корни). Ключевик и страница относятся к одной
# категории, если совпадает хотя бы один корень. Так «Страница» в SEO-панели
# помечает ключ всеми городами, где есть страница той же категории.
# Новую услугу (бани, модульные, гаражи…) добавляйте сюда одной строкой.
CATEGORY_RULES = [
    ("каркас",     r"каркас"),
    ("сип",        r"сип"),
    ("сэндвич",    r"сэндвич|сендвич"),
    ("модульн",    r"модульн"),
    ("деревян",    r"деревян|брус|бревен|бревн"),
    ("щитовой",    r"щитов|сборн"),
    ("коттедж",    r"коттедж"),
    ("канадский",  r"канадск"),
    ("быстровозвод", r"быстровозвод"),
    ("баня",       r"бан[яи]|сауна"),
    ("гараж",      r"гараж"),
    ("дача",       r"дачн|\bдача\b"),
]

# Человекочитаемые названия категорий (интентов) для SEO-панели.
CAT_LABEL = {
    "каркас": "Каркасные дома", "сип": "СИП-панельные дома",
    "сэндвич": "Сэндвич-панели", "панел": "Панельные дома",
    "деревян": "Деревянные дома", "модульн": "Модульные дома",
    "щитовой": "Щитовые / сборные", "коттедж": "Коттеджи",
    "канадский": "Канадские дома", "быстровозвод": "Быстровозводимые дома",
    "баня": "Бани / сауны",
    "гараж": "Гаражи", "дача": "Дачные дома",
    "строй": "Стройка под ключ (общее)",
}


def cat_label(cat):
    return CAT_LABEL.get(cat, cat.capitalize())


def categories_of(text):
    """Множество категорий услуги/ключевика по корням из CATEGORY_RULES.
    Доп. «панел» (общепанельные дома) — только если это НЕ сип и НЕ сэндвич,
    чтобы СИП/сэндвич-страницы не примешивались к общим панельным ключам."""
    t = _norm(text)
    cats = {label for label, rx in CATEGORY_RULES if re.search(rx, t)}
    if re.search(r"панел", t) and not (cats & {"сип", "сэндвич"}):
        cats.add("панел")
    # «строй» — общая категория «построить/строительство дома» без привязки к
    # материалу; добавляется, ТОЛЬКО если никакой материал не распознан, чтобы
    # «строительство каркасных домов» осталось в каркасе, а «построить дом»,
    # «стройка под ключ», «застройщики» попали в общую стройку.
    if not cats and re.search(r"постро|застро|строй|строит", t):
        cats.add("строй")
    return cats


# ---------------- OpenAI ----------------
def api_key():
    k = os.environ.get("OPENAI_API_KEY")
    if not k:
        for p in (os.path.join(HERE, ".openai_key"),
                  os.path.join(HERE, "..", "generate-article", ".openai_key")):
            if os.path.exists(p):
                return open(p, encoding="utf-8").read().strip()
    return k


CW_SYSTEM = (
 "Ты — топовый директ-маркетинговый SEO-копирайтер строительной компании HotWell.kz "
 "из Казахстана (с 2012 года; собственный завод СИП-панелей и домокомплектов в Алматы; "
 "офисы в Астане и Алматы; строит по всему Казахстану; от 89 000 ₸/м² в черновую; "
 "коробка за 18–45 дней). Твоя задача — адаптировать заголовки посадочной страницы "
 "под конкретный город и услугу так, чтобы она ранжировалась в Google/Яндекс по запросу "
 "и конвертировала максимум посетителей в заявки. Пиши живо, конкретно, на русском."
)


def cw_prompt(city, service, kw):
    extra = (" Доп. ключевые слова/акценты: %s." % kw) if kw else ""
    return (
     'Город: "%s". Услуга/запрос: "%s".%s '
     "Верни СТРОГО валидный JSON для SEO-адаптации лендинга под этот город и услугу:\n{"
     '"slug": str (латиница-дефис, кратко, напр. astana-karkasnye-doma),'
     '"meta_title": str (<=60 симв, обязательно с услугой и городом),'
     '"meta_description": str (140-160 симв: услуга+город, выгода, оффер, призыв),'
     '"keywords": [str] (8-12, включая «услуга город», синонимы, «под ключ», «цена»),'
     '"h1": str (мощный H1 с услугой и городом, до 70 симв),'
     '"h1_accent": str (часть из h1 для выделения цветом — например название услуги),'
     '"hero_eyebrow": str (короткий надзаголовок над H1, напр. «Каркасные дома · Астана»),'
     '"hero_lead": str (1-2 предложения под H1: главная выгода + локальная привязка к городу + оффер)'
     "}\n"
     "Только JSON, без markdown."
    ) % (city, service, extra)


def gen_copy(city, service, kw, key):
    payload = {"model": TEXT_MODEL, "temperature": 0.75, "max_tokens": 1200,
               "response_format": {"type": "json_object"},
               "messages": [{"role": "system", "content": CW_SYSTEM},
                            {"role": "user", "content": cw_prompt(city, service, kw)}]}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request("https://api.openai.com/v1/chat/completions", data=data,
        headers={"Authorization": "Bearer " + key, "Content-Type": "application/json"})
    try:
        with urllib.request.urlopen(req, timeout=180, context=CTX) as r:
            res = json.loads(r.read())
    except urllib.error.HTTPError as e:
        sys.exit("OpenAI %s: %s" % (e.code, e.read().decode("utf-8", "replace")[:500]))
    return json.loads(res["choices"][0]["message"]["content"])


CONTENT_SYSTEM = (
 "Ты — эксперт-прораб и SEO-копирайтер HotWell.kz (Казахстан, с 2012; собственный завод "
 "СИП-панелей и домокомплектов в Алматы; доставка и монтаж по всему Казахстану; от 89 000 ₸/м² "
 "в черновую; коробка за 18–45 дней; гарантия, договор, поэтапная оплата). Пишешь УНИКАЛЬНЫЙ, "
 "полезный текст под конкретный город и услугу — не вода, а конкретика: местная привязка "
 "(климат региона, логистика и доставка домокомплекта в этот город, сроки), почему технология "
 "подходит, что входит, как проходит стройка. Без воды и штампов. На русском. "
 "ВАЖНО: склоняй название города грамматически верно по падежам "
 "(в Астане, в Караганде, в Усть-Каменогорске, из Алматы), не оставляй его в именительном падеже.")

CONTENT_MODEL = os.environ.get("OPENAI_CONTENT_MODEL", "gpt-4o")


def content_prompt(city, service):
    return (
     'Город: "%s". Услуга: "%s". '
     "Сгенерируй УНИКАЛЬНЫЙ контент именно под эту связку город+услуга. Верни СТРОГО валидный JSON:\n{"
     '"seo_h2": str (H2 секции, с услугой и городом, до 70 симв),'
     '"seo_html": str (2-3 абзаца, 250-400 слов, в тегах <p>...</p>; конкретика по городу и услуге: '
     'местная привязка/климат/логистика доставки в %s, почему технология подходит, что входит, сроки, '
     'цена-ориентир от 89 000 ₸/м²; естественные вхождения «%s в %s»; БЕЗ воды и общих фраз),'
     '"faq": [ {"q": str, "a": str}, ... ] (ровно 5 уникальных вопросов-ответов под город и услугу: '
     'цена/сроки/доставка в этот город/зимнее строительство/гарантия и т.п.; ответы 1-3 предложения)'
     "}\nТолько JSON, без markdown."
    ) % (city, service, city, service, city)


def gen_content(city, service, key):
    payload = {"model": CONTENT_MODEL, "temperature": 0.7, "max_tokens": 1600,
               "response_format": {"type": "json_object"},
               "messages": [{"role": "system", "content": CONTENT_SYSTEM},
                            {"role": "user", "content": content_prompt(city, service)}]}
    data = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request("https://api.openai.com/v1/chat/completions", data=data,
        headers={"Authorization": "Bearer " + key, "Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=180, context=CTX) as r:
        res = json.loads(r.read())
    return json.loads(res["choices"][0]["message"]["content"])


def fallback_copy(city, service, kw):
    s = service.strip()
    sc = s[0].upper() + s[1:]
    return {
     "slug": slugify(city + "-" + s),
     "meta_title": ("%s в %s под ключ — HotWell.kz" % (sc, city))[:60],
     "meta_description": ("%s в %s от завода-производителя HotWell.kz: от 89 000 ₸/м², коробка "
        "за 18–45 дней. Бесплатный расчёт сметы — оставьте заявку." % (sc, city))[:158],
     "keywords": [s, "%s %s" % (s, city), "%s под ключ" % s, "%s цена %s" % (s, city),
                  "сип панели %s" % city, "строительство домов %s" % city],
     "h1": "%s в %s под ключ" % (sc, city),
     "h1_accent": s,
     "hero_eyebrow": "%s · %s" % (sc, city),
     "hero_lead": ("Строим %s в %s и по всему Казахстану. Собственный завод, цена от "
                   "производителя — от 89 000 ₸/м². Рассчитаем смету онлайн за 5 минут." % (s, city)),
    }


# ---------------- адаптация index.html ----------------
def localize(html, copy, city, service):
    slug = slugify(copy.get("slug") or (city + "-" + service))
    url = "%s/%s.html" % (BASE_URL, slug)
    mt = esc(copy["meta_title"]); md = esc(copy["meta_description"])
    kw = esc(", ".join(copy.get("keywords", [])))

    def sub1(pat, rep):
        return re.subn(pat, lambda m: rep, html, count=1, flags=re.S)

    # <title>
    html = re.sub(r"<title>.*?</title>", "<title>%s</title>" % mt, html, count=1, flags=re.S)
    # meta description
    html = re.sub(r'(<meta name="description" content=")[^"]*(">)',
                  lambda m: m.group(1) + md + m.group(2), html, count=1)
    # keywords (добавляем после description, если ещё нет)
    if 'name="keywords"' not in html:
        html = re.sub(r'(<meta name="description" content="[^"]*">)',
                      lambda m: m.group(1) + '\n<meta name="keywords" content="%s">' % kw, html, count=1)
    # canonical
    html = re.sub(r'(<link rel="canonical" href=")[^"]*(">)',
                  lambda m: m.group(1) + esc(url) + m.group(2), html, count=1)
    # OG/Twitter
    for prop, val in (("og:title", mt), ("og:description", md), ("og:url", esc(url))):
        html = re.sub(r'(<meta property="%s" content=")[^"]*(">)' % re.escape(prop),
                      lambda m, v=val: m.group(1) + v + m.group(2), html, count=1)
    for name, val in (("twitter:title", mt), ("twitter:description", md)):
        html = re.sub(r'(<meta name="%s" content=")[^"]*(">)' % re.escape(name),
                      lambda m, v=val: m.group(1) + v + m.group(2), html, count=1)

    # hero eyebrow / h1 / lead
    html = re.sub(r'<span class="eyebrow">Дома из СИП-панелей</span>',
                  '<span class="eyebrow">%s</span>' % esc(copy["hero_eyebrow"]), html, count=1)
    h1_text = copy["h1"]; acc = (copy.get("h1_accent") or "").strip()
    if acc and acc.lower() in h1_text.lower():
        i = h1_text.lower().index(acc.lower())
        h1_html = (esc(h1_text[:i]) + '<span class="accent">' + esc(h1_text[i:i+len(acc)]) +
                   '</span>' + esc(h1_text[i+len(acc):]))
    else:
        h1_html = esc(h1_text)
    html = re.sub(r"<h1>.*?</h1>", "<h1>%s</h1>" % h1_html, html, count=1, flags=re.S)
    html = re.sub(r'<p class="lead">Работаем с 2012 года.*?</p>',
                  '<p class="lead">%s</p>' % esc(copy["hero_lead"]), html, count=1, flags=re.S)

    # метка источника в форму заявки
    src = esc("Лендинг: %s / %s" % (city, service))
    html = re.sub(r'(<input type="hidden" name="form-name" value="lead">)',
                  lambda m: m.group(1) + '\n      <input type="hidden" name="Источник" value="%s">' % src,
                  html, count=1)

    # JSON-LD Service + BreadcrumbList
    ld = {"@context": "https://schema.org", "@graph": [
        {"@type": "Service", "name": "%s в %s" % (service, city), "serviceType": service,
         "areaServed": {"@type": "City", "name": city},
         "provider": {"@type": "Organization", "name": "HotWell.kz", "url": BASE_URL + "/"},
         "url": url, "description": copy["meta_description"]},
        {"@type": "BreadcrumbList", "itemListElement": [
            {"@type": "ListItem", "position": 1, "name": "Главная", "item": BASE_URL + "/"},
            {"@type": "ListItem", "position": 2, "name": copy["h1"], "item": url}]}]}
    tag = '<script type="application/ld+json">%s</script>\n</head>' % json.dumps(ld, ensure_ascii=False)
    html = html.replace("</head>", tag, 1)

    # УНИКАЛЬНЫЙ контент (текст + FAQ) — снимает риск дубль-фильтра Google.
    # Заполняем маркер UNIQ; если у страницы нет seo_html — маркер остаётся пустым.
    seo_html = (copy.get("seo_html") or "").strip()
    faq = copy.get("faq") or []
    if seo_html or faq:
        h2 = esc(copy.get("seo_h2") or ("%s в %s" % (service, city)))
        block = '<section class="section uniq"><div class="container"><div class="uniq__text"><h2>%s</h2>%s</div>' % (h2, seo_html)
        if faq:
            items = "".join(
                '<details><summary>%s</summary><p>%s</p></details>' % (esc(q.get("q","")), esc(q.get("a","")))
                for q in faq)
            block += '<div class="uniq__faq"><h3>Частые вопросы — %s</h3>%s</div>' % (esc(city), items)
            faq_ld = {"@context": "https://schema.org", "@type": "FAQPage", "mainEntity": [
                {"@type": "Question", "name": q.get("q",""),
                 "acceptedAnswer": {"@type": "Answer", "text": q.get("a","")}} for q in faq]}
            block += '<script type="application/ld+json">%s</script>' % json.dumps(faq_ld, ensure_ascii=False)
        block += '</div></section>'
        html = re.sub(r'<!-- UNIQ:START -->.*?<!-- UNIQ:END -->',
                      lambda m: '<!-- UNIQ:START -->%s<!-- UNIQ:END -->' % block, html, count=1, flags=re.S)

    # маркеры CITYSEL/CITYSVC/CITYNAV уже есть в index.html — клон их наследует,
    # rebuild_city_blocks() заполняет их актуальным содержимым.

    return slug, url, html


def rebuild_city_blocks():
    """Перестраивает меню выбора города (топбар) и блок «Услуги в г. X» (футер)
    на ВСЕХ лендингах по реестру published.json. Так новая страница города
    автоматически появляется в футере этого города и в меню."""
    if not os.path.exists(PUBLISHED):
        return
    pub = json.load(open(PUBLISHED, encoding="utf-8"))
    cities, main_of, by_city = [], {}, {}
    for p in pub:
        c = p["city"]
        if c not in cities:
            cities.append(c)
        main_of.setdefault(c, p["slug"])
        by_city.setdefault(c, []).append(p)
    # цели: все лендинги + главная (у главной нет «текущего» города)
    targets = [(p["slug"] + ".html", p["city"]) for p in pub]
    targets.append(("index.html", None))
    def svc_text(q):
        s = (q.get("service") or "").strip()
        return s[:1].upper() + s[1:] if s else q.get("h1", "")

    for fname, city in targets:
        path = os.path.join(SITE, fname)
        if not os.path.exists(path):
            continue
        html = open(path, encoding="utf-8").read()
        # 1) топбар: меню города убрано — очищаем маркеры
        html = re.sub(r'<!-- CITYSEL:START -->.*?<!-- CITYSEL:END -->',
                      '<!-- CITYSEL:START --><!-- CITYSEL:END -->', html, flags=re.S)
        # 2) колонка «Услуги» в футере: на лендинге — услуги ЭТОГО города;
        #    на главной колонку не показываем (страниц много — перечислять незачем)
        if city and city in by_city:
            svc = "".join('<a href="%s.html">%s</a>' % (esc(q["slug"]), esc(svc_text(q))) for q in by_city[city])
            svcblock = '<div><h4>Услуги в г. %s</h4>%s</div>' % (esc(city), svc)
        else:
            svcblock = ''
        if '<!-- CITYSVC:START -->' in html:
            html = re.sub(r'<!-- CITYSVC:START -->.*?<!-- CITYSVC:END -->',
                          '<!-- CITYSVC:START -->%s<!-- CITYSVC:END -->' % svcblock, html, flags=re.S)
        # 3) «Города» (переключатель) — компактный сворачиваемый <details>.
        #    Ссылки остаются в HTML (внутренняя перелинковка/SEO), но визуально
        #    свёрнуты в одну строку и раскрываются сеткой в 2–3 колонки.
        city_links = "".join('<a%s href="%s.html">%s</a>' % (' class="is-current"' if c == city else '',
                             esc(main_of[c]), esc(c)) for c in sorted(cities))
        summary = ('Города: <b>%s</b>' % esc(city)) if city else 'Выберите город'
        nav = ('<details class="footer-city"><summary>%s</summary>'
               '<div class="footer-city__links">%s</div></details>' % (summary, city_links))
        if '<!-- CITYNAV:START -->' in html:
            html = re.sub(r'<!-- CITYNAV:START -->.*?<!-- CITYNAV:END -->',
                          '<!-- CITYNAV:START -->%s<!-- CITYNAV:END -->' % nav, html, flags=re.S)
        open(path, "w", encoding="utf-8").write(html)


def update_sitemap(url):
    if not os.path.exists(SITEMAP):
        return
    sm = open(SITEMAP, encoding="utf-8").read()
    if url in sm:
        return
    sm = sm.replace("</urlset>", "  <url><loc>%s</loc></url>\n</urlset>" % esc(url))
    open(SITEMAP, "w", encoding="utf-8").write(sm)


def sync_all():
    """Пересобирает ВСЕ лендинги из ТЕКУЩЕЙ главной (site/index.html), сохраняя их
    SEO-данные из published.json. Запускать после изменения блоков на главной —
    правки разойдутся по всем городам/страницам без обращения к OpenAI."""
    if not os.path.exists(PUBLISHED):
        print("Нет published.json — лендингов ещё нет."); return 0
    pub = json.load(open(PUBLISHED, encoding="utf-8"))
    base = open(INDEX, encoding="utf-8").read()
    n = 0
    for p in pub:
        if not p.get("meta_title"):
            print("  ⚠ %s: нет сохранённого SEO — пропуск (перегенерируйте страницу)" % p.get("slug"))
            continue
        slug, url, html = localize(base, p, p["city"], p["service"])
        open(os.path.join(SITE, slug + ".html"), "w", encoding="utf-8").write(html)
        update_sitemap(url)
        n += 1
    rebuild_city_blocks()
    return n


DASHBOARD_FILE = "seo-panel.html"

DASH_CSS = """*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Manrope','Segoe UI',Roboto,Arial,sans-serif;background:#f4f1ea;color:#2E3823;padding:24px}
.wrap{max-width:1240px;margin:0 auto}
h1{font-size:1.7rem}.sub{color:#6b6f5e;margin:4px 0 18px;font-size:.95rem}
.note{background:#fff4d6;border:1px solid #f0d98a;color:#7a5b00;border-radius:10px;padding:10px 14px;font-size:.86rem;margin-bottom:18px}
.stats{display:flex;flex-wrap:wrap;gap:14px;margin-bottom:22px}
.stat{background:#fff;border:1px solid #e2ddd0;border-radius:14px;padding:14px 20px;min-width:150px}
.stat b{display:block;font-size:1.7rem;color:#5B6E3A;line-height:1}
.stat span{color:#6b6f5e;font-size:.85rem}
h2{font-size:1.2rem;margin:26px 0 12px}
.toolbar{display:flex;gap:10px;flex-wrap:wrap;margin-bottom:12px}
.toolbar input{flex:1;min-width:220px;padding:11px 14px;border:1px solid #d8d2c4;border-radius:10px;font:inherit}
.tbl-wrap{background:#fff;border:1px solid #e2ddd0;border-radius:14px;overflow:auto;max-height:70vh}
table{border-collapse:collapse;width:100%;font-size:.9rem}
th,td{padding:9px 13px;text-align:left;border-bottom:1px solid #eee7da;white-space:nowrap}
th{position:sticky;top:0;background:#2E3823;color:#fff;cursor:pointer;user-select:none;font-weight:700;z-index:1}
th:hover{background:#3a4a2c}td.kw{white-space:normal;min-width:240px;font-weight:600;cursor:pointer;-webkit-user-select:none;user-select:none;-webkit-touch-callout:none;transition:background .15s,box-shadow .15s}
td.kw:active{background:#eef2e2}
td.kw.copied{background:#d6f0d0;box-shadow:inset 0 0 0 2px #5B6E3A}
tbody tr:hover{background:#f8f6ef}
.hint{color:#6b6f5e;font-size:.82rem;margin:-4px 0 12px}
#cptoast{position:fixed;left:50%;bottom:26px;transform:translateX(-50%) translateY(20px);background:#2E3823;color:#fff;padding:11px 18px;border-radius:999px;font-size:.9rem;font-weight:700;box-shadow:0 8px 24px rgba(0,0,0,.25);opacity:0;pointer-events:none;transition:opacity .2s,transform .2s;z-index:50;max-width:90vw;white-space:nowrap;overflow:hidden;text-overflow:ellipsis}
#cptoast.show{opacity:1;transform:translateX(-50%) translateY(0)}
.badge{display:inline-block;padding:3px 10px;border-radius:999px;font-weight:800;font-size:.82rem}
.v-hot{background:#d6f0d0;color:#1f6b1a}.v-high{background:#d3e7f5;color:#16557e}.v-mid{background:#fbe7c2;color:#8a5a00}.v-low{background:#eceadf;color:#6b6f5e}
.c-high{color:#c0392b;font-weight:800}.c-mid{color:#b8860b;font-weight:800}.c-low{color:#1f7a3a;font-weight:800}
.cost{color:#2E3823;font-weight:700}
.has-page{color:#1f7a3a;font-weight:700}.no-page{color:#b9b4a6}.no-cov{color:#c0392b;font-weight:700}
td.cat{white-space:normal;min-width:150px}
.cat-chip{display:inline-block;background:#E3E9D5;color:#3f5226;border-radius:999px;padding:2px 9px;font-size:.78rem;font-weight:700;margin:1px 3px 1px 0}
.no-cat{color:#b9b4a6;font-size:.82rem}
#it td:first-child{font-weight:700}
.pagelink a{color:#5B6E3A;font-weight:700}
.done{display:inline-block;background:#5B6E3A;color:#fff;border-radius:999px;padding:2px 9px;font-size:.78rem;font-weight:800}
@media(max-width:640px){body{padding:14px}h1{font-size:1.35rem}}"""

DASH_JS = """function flt(id,inp){var q=inp.value.toLowerCase();var rows=document.querySelectorAll('#'+id+' tbody tr');rows.forEach(function(r){r.style.display=r.getAttribute('data-s').indexOf(q)>-1?'':'none'})}
function srt(id,col,num){var tb=document.querySelector('#'+id+' tbody');var rows=[].slice.call(tb.querySelectorAll('tr'));var d=tb.getAttribute('data-d'+col)==='1';tb.setAttribute('data-d'+col,d?'0':'1');rows.sort(function(a,b){var x=a.children[col].getAttribute('data-v')||a.children[col].textContent;var y=b.children[col].getAttribute('data-v')||b.children[col].textContent;if(num){x=parseFloat(x)||0;y=parseFloat(y)||0;return d?x-y:y-x}return d?(''+y).localeCompare(x):(''+x).localeCompare(y)});rows.forEach(function(r){tb.appendChild(r)})}
function cpFallback(t){var ta=document.createElement('textarea');ta.value=t;ta.style.position='fixed';ta.style.top='-1000px';ta.style.opacity='0';document.body.appendChild(ta);ta.focus();ta.select();try{document.execCommand('copy')}catch(e){}document.body.removeChild(ta);}
function cpText(t){try{if(navigator.clipboard&&navigator.clipboard.writeText){navigator.clipboard.writeText(t)['catch'](function(){cpFallback(t)});return;}}catch(e){}cpFallback(t);}
function cpToast(msg){var el=document.getElementById('cptoast');if(!el)return;el.textContent=msg;el.classList.add('show');clearTimeout(el._t);el._t=setTimeout(function(){el.classList.remove('show')},1500);}
function cpFire(c){var txt=(c.textContent||'').trim();if(!txt)return;cpText(txt);c.classList.add('copied');setTimeout(function(){c.classList.remove('copied')},650);cpToast('Скопировано: '+txt);if(navigator.vibrate){try{navigator.vibrate(30)}catch(e){}}}
function cpInit(){var cells=document.querySelectorAll('td.kw');cells.forEach(function(c){var timer=null;var done=false;function start(){done=false;timer=setTimeout(function(){done=true;cpFire(c)},450);}function cancel(){if(timer){clearTimeout(timer);timer=null;}}c.addEventListener('touchstart',start,{passive:true});c.addEventListener('touchend',function(e){if(done){e.preventDefault();}cancel();});c.addEventListener('touchmove',cancel,{passive:true});c.addEventListener('touchcancel',cancel);c.addEventListener('mousedown',start);c.addEventListener('mouseup',cancel);c.addEventListener('mouseleave',cancel);c.addEventListener('contextmenu',function(e){e.preventDefault();});});}
document.addEventListener('DOMContentLoaded',cpInit);"""


def _vol_class(v):
    try: v = int(v)
    except Exception: return "v-low", 0
    if v >= 1000: return "v-hot", v
    if v >= 300: return "v-high", v
    if v >= 100: return "v-mid", v
    return "v-low", v


def build_dashboard():
    """Строит внутреннюю SEO-панель site/seo-panel.html (noindex):
    таблица ключей из Лист2 (объём/конкуренция/цена, цветовая кодировка) +
    таблица созданных страниц по городам. Источники: Ключевые слова.xlsx + published.json."""
    kws = []
    try:
        import openpyxl
        if os.path.exists(KEYWORDS_XLSX):
            wb = openpyxl.load_workbook(KEYWORDS_XLSX, data_only=True)
            if "Лист2" in wb.sheetnames:
                ws = wb["Лист2"]
                for r in range(2, ws.max_row + 1):
                    kw = ws.cell(r, 1).value
                    if not kw:
                        continue
                    kws.append({"kw": str(kw).strip(), "vol": ws.cell(r, 3).value or 0,
                                "comp": (ws.cell(r, 6).value or "").strip(),
                                "lo": ws.cell(r, 8).value, "hi": ws.cell(r, 9).value})
    except ImportError:
        print("  (openpyxl не установлен — таблица ключей в панель не попадёт)")
    pub = json.load(open(PUBLISHED, encoding="utf-8")) if os.path.exists(PUBLISHED) else []

    # Категория каждой страницы → города этой категории.
    cat_cities = {}
    for p in pub:
        for cat in categories_of(p.get("service") or ""):
            cat_cities.setdefault(cat, set()).add(p["city"])

    def cities_for(kw):
        out = set()
        for cat in categories_of(kw):
            out |= cat_cities.get(cat, set())
        return sorted(out)

    kws.sort(key=lambda k: _vol_class(k["vol"])[1], reverse=True)
    comp_cls = {"Высокий": "c-high", "Средний": "c-mid", "Низкий": "c-low"}

    # Агрегат по интентам (категориям): ключей, объём/мес, покрытие городами.
    intents = {}
    for k in kws:
        vn = _vol_class(k["vol"])[1]
        cats = sorted(categories_of(k["kw"])) or ["__none__"]
        k["_cats"] = cats
        for c in cats:
            d = intents.setdefault(c, {"keys": 0, "vol": 0})
            d["keys"] += 1; d["vol"] += vn
    for c in cat_cities:                     # категории, под которые есть страница, но нет ключа
        intents.setdefault(c, {"keys": 0, "vol": 0})

    def intent_label(c):
        return "Прочее (без категории)" if c == "__none__" else cat_label(c)

    irows = []
    for c, d in sorted(intents.items(), key=lambda x: (x[1]["vol"], x[1]["keys"]), reverse=True):
        cities = sorted(cat_cities.get(c, set()))
        cov = ('<span class="has-page">✓ %s</span>' % ", ".join(esc(x) for x in cities)) if cities \
              else '<span class="no-cov">нет страницы</span>'
        irows.append(
            '<tr data-s="%s"><td class="cat"><span class="cat-chip">%s</span></td>'
            '<td data-v="%d">%d</td><td data-v="%d">%s</td><td>%s</td></tr>'
            % (esc(intent_label(c).lower()), esc(intent_label(c)),
               d["keys"], d["keys"], d["vol"], "{:,}".format(d["vol"]).replace(",", " ") if d["vol"] else "—", cov))

    krows = []
    for i, k in enumerate(kws, 1):
        vc, vn = _vol_class(k["vol"])
        cc = comp_cls.get(k["comp"], "")
        cost = ""
        if k["lo"] is not None and k["hi"] is not None:
            try: cost = "$%.2f–$%.2f" % (float(k["lo"]), float(k["hi"]))
            except Exception: cost = ""
        cities = cities_for(k["kw"])
        page = ('<span class="has-page">✓ %s</span>' % ", ".join(esc(c) for c in cities)) if cities else '<span class="no-page">—</span>'
        cats = [c for c in k["_cats"] if c != "__none__"]
        intent_html = ("".join('<span class="cat-chip">%s</span>' % esc(cat_label(c)) for c in cats)
                       if cats else '<span class="no-cat">прочее</span>')
        intent_search = " ".join(intent_label(c) for c in k["_cats"]).lower()
        krows.append(
            '<tr data-s="%s"><td>%d</td><td class="kw">%s</td><td class="cat">%s</td>'
            '<td data-v="%d"><span class="badge %s">%s</span></td>'
            '<td class="%s">%s</td><td class="cost">%s</td><td>%s</td></tr>'
            % (esc((k["kw"] + " " + intent_search).lower()), i, esc(k["kw"]), intent_html,
               vn, vc, ("%d" % vn if vn else "—"),
               cc, esc(k["comp"] or "—"), cost or "—", page))

    prows = []
    for p in sorted(pub, key=lambda x: (x["city"], x["service"])):
        prows.append('<tr data-s="%s"><td>%s</td><td>%s</td><td>%s</td>'
                     '<td class="pagelink"><a href="%s" target="_blank" rel="noopener">%s</a></td>'
                     '<td>%s</td><td><span class="done">готово</span></td></tr>'
                     % (esc((p["city"] + " " + p["service"]).lower()), esc(p["city"]), esc(p["service"]),
                        esc(p.get("h1", "")), esc(p["url"]), esc(p["url"].replace("https://", "")),
                        esc(p.get("Создано", p.get("date", "")))))

    total_vol = sum(_vol_class(k["vol"])[1] for k in kws)
    html = (
'<!DOCTYPE html><html lang="ru"><head><meta charset="UTF-8">'
'<meta name="viewport" content="width=device-width, initial-scale=1">'
'<meta name="robots" content="noindex, nofollow">'
'<title>SEO-панель · HotWell.kz (внутренняя)</title>'
'<style>' + DASH_CSS + '</style></head><body><div class="wrap">'
'<h1>SEO-панель HotWell.kz</h1>'
'<div class="sub">Семантика и созданные посадочные страницы по городам</div>'
'<div class="note">🔒 Внутренняя страница — закрыта от индексации (noindex) и не в карте сайта. '
'Ссылку никому не передавайте — доступ по прямому URL.</div>'
'<div class="stats">'
'<div class="stat"><b>%d</b><span>ключевых слов</span></div>'
'<div class="stat"><b>%s</b><span>суммарный объём/мес</span></div>'
'<div class="stat"><b>%d</b><span>создано страниц</span></div>'
'<div class="stat"><b>%d</b><span>городов</span></div>'
'</div>'
% (len(kws), "{:,}".format(total_vol).replace(",", " "), len(pub),
   len(set(p["city"] for p in pub)))
+ '<h2>Интенты услуг (группы ключей)</h2>'
'<div class="hint">Один интент = одна страница на город. Если у интента «нет страницы», '
'а объём большой — это незакрытая возможность. Делать отдельный лендинг под каждый '
'синоним ОДНОГО интента не нужно (каннибализация) — хватает одной страницы на город.</div>'
'<div class="tbl-wrap"><table id="it"><thead><tr>'
'<th onclick="srt(\'it\',0,0)">Интент / категория</th><th onclick="srt(\'it\',1,1)">Ключей</th>'
'<th onclick="srt(\'it\',2,1)">Объём/мес</th><th onclick="srt(\'it\',3,0)">Покрытие (города)</th>'
'</tr></thead><tbody>' + "".join(irows) + '</tbody></table></div>'
+ '<h2>Ключевые слова</h2>'
'<div class="hint">💡 Удерживайте ключевое слово (долгий тап / зажатие мышью) — оно скопируется в буфер обмена. '
'Колонка «Интент» показывает, к какой странице относится ключ.</div>'
'<div class="toolbar"><input type="search" placeholder="Поиск по ключу или интенту…" oninput="flt(\'kt\',this)"></div>'
'<div class="tbl-wrap"><table id="kt"><thead><tr>'
'<th onclick="srt(\'kt\',0,1)">#</th><th onclick="srt(\'kt\',1,0)">Ключевое слово</th>'
'<th onclick="srt(\'kt\',2,0)">Интент</th>'
'<th onclick="srt(\'kt\',3,1)">Запросов/мес</th><th onclick="srt(\'kt\',4,0)">Конкуренция</th>'
'<th onclick="srt(\'kt\',5,0)">Цена клика</th><th onclick="srt(\'kt\',6,0)">Страница</th>'
'</tr></thead><tbody>' + "".join(krows) + '</tbody></table></div>'
'<h2>Созданные страницы по городам</h2>'
'<div class="toolbar"><input type="search" placeholder="Поиск по городу/услуге…" oninput="flt(\'pt\',this)"></div>'
'<div class="tbl-wrap"><table id="pt"><thead><tr>'
'<th onclick="srt(\'pt\',0,0)">Город</th><th onclick="srt(\'pt\',1,0)">Ключевое слово / услуга</th>'
'<th>Заголовок (H1)</th><th>Ссылка</th><th onclick="srt(\'pt\',4,0)">Создано</th><th>Статус</th>'
'</tr></thead><tbody>' + "".join(prows) + '</tbody></table></div>'
'</div><div id="cptoast"></div>'
'<script>' + DASH_JS + '</script></body></html>')
    open(os.path.join(SITE, DASHBOARD_FILE), "w", encoding="utf-8").write(html)
    print("  • SEO-панель обновлена: site/%s (ключей: %d, страниц: %d)" % (DASHBOARD_FILE, len(kws), len(pub)))


def update_keywords_xlsx(city, service, h1, url):
    """Ведёт лист по каждому городу в «Ключевые слова.xlsx»: создаёт лист города,
    если его нет, и записывает строку «ключевое слово/услуга → ссылка на страницу».
    Так видно, под какой запрос какая страница уже сделана."""
    try:
        import openpyxl
    except ImportError:
        print("  (openpyxl не установлен — лист города в xlsx не обновлён)")
        return
    if not os.path.exists(KEYWORDS_XLSX):
        print("  (Ключевые слова.xlsx не найден — пропуск)")
        return
    import datetime
    today = datetime.date.today().isoformat()
    wb = openpyxl.load_workbook(KEYWORDS_XLSX)
    title = city.strip()[:31]
    if title in wb.sheetnames:
        ws = wb[title]
    else:
        ws = wb.create_sheet(title)
        ws.append(["Ключевое слово / услуга", "Заголовок (H1)", "Ссылка", "Создано"])
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 48
    row_idx = None
    for r in range(2, ws.max_row + 1):
        if (ws.cell(r, 3).value or "") == url:
            row_idx = r; break
    vals = [service, h1, url, today]
    if row_idx is None:
        ws.append(vals); row_idx = ws.max_row
    else:
        for i, v in enumerate(vals, 1):
            ws.cell(row_idx, i).value = v
    c = ws.cell(row_idx, 3); c.hyperlink = url; c.style = "Hyperlink"
    wb.save(KEYWORDS_XLSX)
    print("  • «Ключевые слова.xlsx» → лист «%s»: %s" % (title, service))


def backfill_content(limit=0):
    """Догенерирует уникальный контент (seo_html + FAQ) для всех опубликованных
    лендингов, у которых его ещё нет, сохраняет в published.json и пересобирает
    все страницы из текущей главной. Снимает риск дубль-фильтра Google."""
    if not os.path.exists(PUBLISHED):
        sys.exit("Нет published.json")
    key = api_key()
    if not key:
        sys.exit("OPENAI_API_KEY не найден — бэкафилл невозможен.")
    pub = json.load(open(PUBLISHED, encoding="utf-8"))
    todo = [p for p in pub if not (p.get("seo_html") or "").strip()]
    if limit:
        todo = todo[:limit]
    print("Бэкафилл уникального контента (%s): %d из %d страниц…" % (CONTENT_MODEL, len(todo), len(pub)))
    done = 0
    for i, p in enumerate(todo, 1):
        try:
            c = gen_content(p["city"], p["service"], key)
            p["seo_h2"] = c.get("seo_h2", ""); p["seo_html"] = c.get("seo_html", ""); p["faq"] = c.get("faq", [])
            done += 1
            print("  [%d/%d] %s — ok" % (i, len(todo), p["slug"]))
        except Exception as e:
            print("  [%d/%d] %s — ОШИБКА: %s" % (i, len(todo), p["slug"], e))
        if i % 20 == 0:                                  # периодически сохраняем прогресс
            json.dump(pub, open(PUBLISHED, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    json.dump(pub, open(PUBLISHED, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    print("Сгенерировано: %d. Пересобираю страницы из главной…" % done)
    n = sync_all()
    build_dashboard()
    print("✓ Готово: контент добавлен, пересобрано %d лендингов." % n)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--city")
    ap.add_argument("--service", help="услуга или ключевые слова")
    ap.add_argument("--kw", default="")
    ap.add_argument("--no-ai", action="store_true")
    ap.add_argument("--sync", action="store_true",
                    help="пересобрать все лендинги из текущей главной (после правок блоков)")
    ap.add_argument("--dashboard", action="store_true",
                    help="только пересобрать SEO-панель (site/seo-panel.html)")
    ap.add_argument("--backfill", action="store_true",
                    help="догенерировать уникальный контент+FAQ для всех лендингов без него и пересобрать")
    ap.add_argument("--limit", type=int, default=0, help="ограничить число страниц на бэкафилл (0 = все)")
    args = ap.parse_args()

    if not os.path.exists(INDEX):
        sys.exit("Не найден site/index.html")

    if args.dashboard:
        build_dashboard()
        return

    if args.backfill:
        backfill_content(limit=args.limit)
        return

    if args.sync:
        print("→ Синхронизирую все лендинги с текущей главной…")
        n = sync_all()
        build_dashboard()
        print("\n✓ Готово: пересобрано лендингов: %d (правки главной разнесены по всем страницам)." % n)
        return

    if not args.city or not args.service:
        sys.exit("Укажите --city и --service (или --sync для синхронизации с главной).")

    if args.no_ai:
        copy = fallback_copy(args.city, args.service, args.kw)
        print("• SEO-тексты: шаблон (--no-ai)")
    else:
        key = api_key()
        if not key:
            print("⚠ OPENAI_API_KEY не найден — использую шаблон.")
            copy = fallback_copy(args.city, args.service, args.kw)
        else:
            print("→ Адаптирую SEO под «%s» в «%s» (OpenAI)…" % (args.service, args.city))
            copy = gen_copy(args.city, args.service, args.kw, key)
            try:
                print("→ Генерирую уникальный контент + FAQ (%s)…" % CONTENT_MODEL)
                copy.update(gen_content(args.city, args.service, key))
            except Exception as e:
                print("  ⚠ уникальный контент не сгенерирован (%s) — страница без него" % e)

    base_html = open(INDEX, encoding="utf-8").read()
    slug, url, html = localize(base_html, copy, args.city, args.service)
    path = os.path.join(SITE, slug + ".html")
    open(path, "w", encoding="utf-8").write(html)
    update_sitemap(url)

    pub = []
    if os.path.exists(PUBLISHED):
        try: pub = json.load(open(PUBLISHED, encoding="utf-8"))
        except Exception: pub = []
    entry = dict(copy)                              # сохраняем все SEO-поля (для --sync)
    entry.update({"city": args.city, "service": args.service, "slug": slug, "url": url})
    pub = [p for p in pub if p.get("slug") != slug]
    pub.append(entry)
    json.dump(pub, open(PUBLISHED, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

    # обновляем меню города и футер по городу на всех лендингах
    rebuild_city_blocks()
    # ведём реестр ключевиков по городам в xlsx + перестраиваем SEO-панель
    update_keywords_xlsx(args.city, args.service, copy.get("h1", ""), url)
    build_dashboard()

    print("\n✓ Готово (полная копия главной, SEO под город+услугу):")
    print("  • Страница: site/%s.html" % slug)
    print("  • URL:      %s" % url)
    print("  • H1:       %s" % copy["h1"])
    print("  • Добавлен в sitemap.xml")


if __name__ == "__main__":
    main()
