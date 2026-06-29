#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Дорабатывает заполненный менеджером Excel:
- добавляет колонку «Готово» (отмечает уже просмотренные строки),
- убирает «Построенные объекты»,
- делает «Этажность» редактируемой (дропдаун),
- сохраняет все введённые значения.

Запуск: python3 refine_params_file.py <вход.xlsx> <выход.xlsx> [done_until_slug]
"""
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

IN, OUT = sys.argv[1], sys.argv[2]
DONE_UNTIL = sys.argv[3] if len(sys.argv) > 3 else "m-136"  # слаг-граница (включительно)

src = load_workbook(IN).active
# исходные колонки: 1 слаг,2 ссылка,3 назв,4 раздел,5 площадь,6 этажность,
# 7 спальни,8 габариты,9 h1,10 h2,11 крыша,12 форма,13 цена,14 авто
rows = []
for r in range(2, src.max_row + 1):
    vals = [src.cell(row=r, column=c).value for c in range(1, 15)]
    if vals[0]:
        rows.append(vals)

# граница «готово»: все строки до первой, чей слаг начинается с DONE_UNTIL (включительно)
done = set()
for v in rows:
    done.add(v[0])
    if str(v[0]).startswith(DONE_UNTIL):
        break

FLOORS = ["1 этаж", "1,5 этажа", "2 этажа"]   # 1,5 = мансардный
HEIGHTS = ["2,5 м", "2,8 м", "2,9 м", "3,0 м", "3,5 м", "4,0 м"]
HEIGHTS2 = ["—"] + HEIGHTS
ROOFS = ["1-скатная", "2-скатная", "4-скатная"]
SHAPES = ["Простая", "Сложная"]


def norm_floors(v):
    s = str(v or "").lower()
    if "1.5" in s or "1,5" in s or "мансард" in s:
        return "1,5 этажа"
    if "2" in s or "3" in s:   # 3-этажные сводим к «2 этажа» — оставляем только 3 варианта
        return "2 этажа"
    return "1 этаж"


wb = Workbook()
ws = wb.active
ws.title = "Проекты"
HEAD = ["Готово", "Слаг (НЕ менять)", "Ссылка на проект", "Название", "Раздел",
        "Площадь, м²", "Этажность", "Спальни", "Габариты", "Высота 1 этажа",
        "Высота 2 этажа", "Тип крыши", "Форма дома", "Цена, ₸", "Цена авто?"]
EDIT_COLS = {7, 10, 11, 12, 13}   # этажность, высоты, крыша, форма

hfill = PatternFill("solid", fgColor="1A7D3C"); hfont = Font(bold=True, color="FFFFFF", size=10)
editfill = PatternFill("solid", fgColor="FFF6D5")
donefill = PatternFill("solid", fgColor="D6EAD9")
okfill = PatternFill("solid", fgColor="1A7D3C")
thin = Side(style="thin", color="DDDDDD"); border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEAD)
for ci in range(1, len(HEAD) + 1):
    c = ws.cell(row=1, column=ci); c.fill = hfill; c.font = hfont
    c.alignment = Alignment(vertical="center", wrap_text=True); c.border = border
ws.row_dimensions[1].height = 30

kept = 0; removed = 0; done_marked = 0
for v in rows:
    if v[3] == "Построенные объекты":   # раздел
        removed += 1
        continue
    kept += 1
    is_done = v[0] in done
    if is_done:
        done_marked += 1
    # Готово, слаг, ссылка, назв, раздел, площадь, этажность(норм), спальни, габариты,
    # h1, h2, крыша, форма, цена, авто
    ws.append(["✓" if is_done else "", v[0], v[1], v[2], v[3], v[4], norm_floors(v[5]),
               v[6], v[7], v[8], v[9], v[10], v[11], v[12], v[13]])

nrows = ws.max_row
for r in range(2, nrows + 1):
    is_done = ws.cell(row=r, column=1).value == "✓"
    for ci in range(1, len(HEAD) + 1):
        cell = ws.cell(row=r, column=ci); cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(ci == 4),
                                   horizontal=("center" if ci == 1 else "general"))
        if ci == 1 and is_done:
            cell.fill = okfill; cell.font = Font(bold=True, color="FFFFFF")
        elif ci in EDIT_COLS:
            cell.fill = editfill
        elif is_done and ci in (2, 3, 4, 5):
            cell.fill = donefill


# Значения с запятыми («1,5 этажа», «2,5 м») нельзя задавать инлайн (запятая —
# разделитель списка). Кладём списки на скрытый лист и ссылаемся диапазоном.
lists = wb.create_sheet("Списки")
ranges = {}
_col = 1
for key, opts in [("floors", FLOORS), ("h1", HEIGHTS), ("h2", HEIGHTS2),
                  ("roof", ROOFS), ("shape", SHAPES)]:
    for i, o in enumerate(opts, 1):
        lists.cell(row=i, column=_col, value=o)
    cl = get_column_letter(_col)
    ranges[key] = "Списки!$%s$1:$%s$%d" % (cl, cl, len(opts))
    _col += 1
lists.sheet_state = "hidden"


def add_dv(col_idx, ref, allow_blank=True):
    dv = DataValidation(type="list", formula1=ref, allow_blank=allow_blank, showErrorMessage=True)
    dv.error = "Выберите значение из списка"; dv.prompt = "Выберите из списка"
    ws.add_data_validation(dv)
    col = get_column_letter(col_idx); dv.add("%s2:%s%d" % (col, col, nrows))


_ok = DataValidation(type="list", formula1='"✓"', allow_blank=True)
ws.add_data_validation(_ok); _ok.add("A2:A%d" % nrows)
add_dv(7, ranges["floors"], allow_blank=False)
add_dv(10, ranges["h1"])
add_dv(11, ranges["h2"])
add_dv(12, ranges["roof"])
add_dv(13, ranges["shape"])

widths = [9, 26, 44, 30, 20, 11, 13, 9, 13, 14, 14, 15, 13, 14, 11]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "B2"
ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(HEAD)), nrows)
wb.save(OUT)
print("Готово:", OUT)
print("Оставлено проектов:", kept, "| убрано построенных:", removed, "| помечено готово:", done_marked)
