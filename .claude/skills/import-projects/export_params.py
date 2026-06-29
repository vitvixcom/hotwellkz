#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Выгрузка всех проектов в Excel для ручного заполнения недостающих параметров
расчёта (высота 1/2 этажа, тип крыши, форма дома). Колонки с дропдаунами,
редактируемые поля подсвечены. Слаг — ключ для обратной загрузки (НЕ менять).

Запуск:  python3 export_params.py [путь_вывода.xlsx]
"""
import os, sys, re, importlib.util
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
BASE_URL = os.environ.get("SITE_BASE_URL", "https://hotwellkz.kz").rstrip("/")
OUT = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "projects_params.xlsx")

spec = importlib.util.spec_from_file_location("imp", os.path.join(HERE, "import.py"))
imp = importlib.util.module_from_spec(spec)
spec.loader.exec_module(imp)
items = imp.parse_csv(os.path.join(HERE, "source.csv"))

HEIGHTS = ["2,5 м", "2,8 м", "2,9 м", "3,0 м", "3,5 м", "4,0 м"]
HEIGHTS2 = ["—"] + HEIGHTS
ROOFS = ["1-скатная", "2-скатная", "4-скатная"]
SHAPES = ["Простая", "Сложная"]
_HVALS = [2.5, 2.8, 2.9, 3.0, 3.5, 4.0]


def snap_height(v):
    best = min(_HVALS, key=lambda x: abs(x - v))
    return "%s м" % ("%.1f" % best).replace(".", ",")


def parse_heights(h):
    nums = re.findall(r"\d(?:[.,]\d)?", h or "")
    vals = []
    for n in nums:
        try:
            f = float(n.replace(",", "."))
        except ValueError:
            continue
        if 2.0 <= f <= 6.0:
            vals.append(f)
    return vals


def is_built(p):
    return p["group"] == "Построенные объекты"


def default_h1(p):
    hv = parse_heights(p.get("height"))
    if hv:
        return snap_height(hv[0])
    a = p.get("area") or 0
    return "2,8 м" if a > 50 else "2,5 м"


def default_h2(p):
    fl = p.get("floors") or 1
    if fl < 1.5:
        return "—"
    hv = parse_heights(p.get("height"))
    if len(hv) > 1:
        return snap_height(hv[1])
    a = p.get("area") or 0
    if a <= 50:
        return "2,5 м"
    return "2,8 м" if a > 100 else "2,5 м"


wb = Workbook()
ws = wb.active
ws.title = "Проекты"

HEAD = ["Слаг (НЕ менять)", "Ссылка на проект", "Название", "Раздел", "Площадь, м²",
        "Этажность", "Спальни", "Габариты", "Высота 1 этажа", "Высота 2 этажа",
        "Тип крыши", "Форма дома", "Цена, ₸", "Цена авто?"]
EDIT_COLS = {9, 10, 11, 12}  # 1-based: высоты, крыша, форма

hfill = PatternFill("solid", fgColor="1A7D3C")
hfont = Font(bold=True, color="FFFFFF", size=10)
editfill = PatternFill("solid", fgColor="FFF6D5")
thin = Side(style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEAD)
for ci, _ in enumerate(HEAD, 1):
    c = ws.cell(row=1, column=ci)
    c.fill = hfill; c.font = hfont; c.alignment = Alignment(vertical="center", wrap_text=True); c.border = border
ws.row_dimensions[1].height = 30

for p in items:
    area = (int(p["area"]) if p["area"] and p["area"] == int(p["area"]) else p["area"]) if p["area"] else ""
    ws.append([
        p["slug"],
        "%s/proekty/%s.html" % (BASE_URL, p["slug"]),
        p["name"], p["group"], area,
        p.get("floors_txt") or "", p.get("bedrooms") or "", p.get("dims") or "",
        default_h1(p), default_h2(p), "2-скатная", "Простая",
        p["price"] or "", "да" if p.get("price_est") else ("" if p["price"] else "нет цены"),
    ])

nrows = len(items) + 1
# подсветка редактируемых колонок + границы
for r in range(2, nrows + 1):
    for ci in range(1, len(HEAD) + 1):
        cell = ws.cell(row=r, column=ci)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(ci in (3,)))
        if ci in EDIT_COLS:
            cell.fill = editfill

# дропдауны
def add_dv(col_idx, options, allow_blank=True):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(options), allow_blank=allow_blank, showErrorMessage=True)
    dv.error = "Выберите значение из списка"; dv.prompt = "Выберите из списка"
    ws.add_data_validation(dv)
    col = get_column_letter(col_idx)
    dv.add("%s2:%s%d" % (col, col, nrows))

add_dv(9, HEIGHTS)
add_dv(10, HEIGHTS2)
add_dv(11, ROOFS)
add_dv(12, SHAPES)

# ширины колонок
widths = [26, 46, 30, 20, 11, 14, 9, 14, 15, 15, 16, 14, 14, 11]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(HEAD)), nrows)

wb.save(OUT)
print("Готово:", OUT)
print("Проектов:", len(items), "| редактируемые колонки подсвечены жёлтым (высоты, крыша, форма).")
